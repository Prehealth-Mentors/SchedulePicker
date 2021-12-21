# Author: Sam Shenoi
# Description: This file allows users to use an excel file to create the graph
#  It extends the basic graph.py class

# Imports
from openpyxl import load_workbook
from graph import Graph,Node
import re
class ExcelGraph(Graph):
    def readfile(self,filename,contents=None):
        # Open up the workbook
        wb = load_workbook(filename=filename)

        # Get the active sheet
        sheet = wb.active

        # Check to see if we are on the header row
        header_row = True

        # Set up the values we need
        self.num_mentors = 0
        self.peopleHash = {}
        header_dict = dict()
        feature_dict = dict()

        for row in sheet.iter_rows():
            row_values = [cell.value for cell in row]
            # In order to get code to work with our other functions
            # list compression the row
            # If this is the header row, save the header
            if header_row:
                feature_dict,header_row = self.firstrow(row_values)
            else:
                self.dataRow(feature_dict,row_values)



        # Calculate the target group size based on the mentor mentee ratio
        self.group_size = int((len(self.peopleHash.keys()) - self.num_mentors)/self.num_mentors) + 1



        return self.peopleHash




